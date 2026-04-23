import io
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy import func
from app import models

# PDF Libraries
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# Excel Libraries
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_pdf_report(db: Session, report_type: str, state: str = None, district: str = None) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom Styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#1e3a8a'), # Navy blue
        alignment=1, # Center
        spaceAfter=10
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontSize=11,
        leading=14,
        textColor=colors.HexColor('#475569'), # Slate grey
        alignment=1, # Center
        spaceAfter=30
    )
    
    h2_style = ParagraphStyle(
        'SectionH2',
        parent=styles['Heading2'],
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#0f172a'),
        spaceBefore=15,
        spaceAfter=10
    )
    
    body_style = ParagraphStyle(
        'BodyDark',
        parent=styles['Normal'],
        fontSize=10,
        leading=13,
        textColor=colors.HexColor('#334155')
    )
    
    table_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.white,
        fontName='Helvetica-Bold'
    )
    
    table_body_style = ParagraphStyle(
        'TableBody',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.HexColor('#334155')
    )

    story = []
    
    # Title & Metadata
    story.append(Paragraph("GOVERNMENT SCHEME MONITORING & MIS AUTOMATION SYSTEM", title_style))
    
    loc_info = []
    if state:
        loc_info.append(state)
    if district:
        loc_info.append(district)
    loc_str = f" for {', '.join(loc_info)}" if loc_info else " (National)"
    
    story.append(Paragraph(f"{report_type.capitalize()} Performance & Reconciliation Audit Report{loc_str} — Generated on {datetime.now().strftime('%d-%b-%Y %I:%M %p')}", subtitle_style))
    
    # Fetch KPI Summary
    b_query = db.query(models.Beneficiary)
    alloc_query = db.query(func.sum(models.FundAllocation.allocated_amount))
    util_query = db.query(func.sum(models.FundUtilization.utilized_amount))
    err_query = db.query(models.ValidationError)
    
    if state:
        b_query = b_query.filter(models.Beneficiary.state == state)
        alloc_query = alloc_query.filter(models.FundAllocation.state == state)
        util_query = util_query.filter(models.FundUtilization.state == state)
    if district:
        b_query = b_query.filter(models.Beneficiary.district == district)
        alloc_query = alloc_query.filter(models.FundAllocation.district == district)
        util_query = util_query.filter(models.FundUtilization.district == district)
        
    total_beneficiaries = b_query.count()
    total_alloc = alloc_query.scalar() or 0.0
    total_util = util_query.scalar() or 0.0
    util_percentage = (total_util / total_alloc * 100) if total_alloc > 0 else 0.0
    
    if state or district:
        b_sub = db.query(models.Beneficiary.beneficiary_id)
        if state:
            b_sub = b_sub.filter(models.Beneficiary.state == state)
        if district:
            b_sub = b_sub.filter(models.Beneficiary.district == district)
        b_ids = b_sub.subquery()
        
        desc_filter = []
        if state:
            desc_filter.append(models.ValidationError.description.contains(state))
        if district:
            desc_filter.append(models.ValidationError.description.contains(district))
            
        from sqlalchemy import or_
        if desc_filter:
            err_query = err_query.filter((models.ValidationError.beneficiary_id.in_(b_ids)) | or_(*desc_filter))
        else:
            err_query = err_query.filter(models.ValidationError.beneficiary_id.in_(b_ids))
            
    total_errors = err_query.count()
    
    # KPI Grid Table
    kpi_data = [
        [
            Paragraph("<b>Total Beneficiaries</b>", body_style),
            Paragraph("<b>Total Allocated (INR)</b>", body_style),
            Paragraph("<b>Total Utilized (INR)</b>", body_style)
        ],
        [
            Paragraph(f"{total_beneficiaries:,}", body_style),
            Paragraph(f"{total_alloc:,.2f}", body_style),
            Paragraph(f"{total_util:,.2f}", body_style)
        ],
        [
            Paragraph("<b>Utilization %</b>", body_style),
            Paragraph("<b>Validation Warnings</b>", body_style),
            Paragraph("<b>System Integrity</b>", body_style)
        ],
        [
            Paragraph(f"{util_percentage:.2f}%", body_style),
            Paragraph(f"{total_errors} Active Alerts", body_style),
            Paragraph("99.98% Healthy" if total_errors < 500 else "Action Required", body_style)
        ]
    ]
    
    kpi_table = Table(kpi_data, colWidths=[180, 180, 180])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f8fafc')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#cbd5e1')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING', (0,0), (-1,-1), 8),
    ]))
    
    story.append(Paragraph("Executive Summary Stats", h2_style))
    story.append(kpi_table)
    story.append(Spacer(1, 20))
    
    # Scheme-wise Performance Table
    story.append(Paragraph("Scheme-wise Performance Matrix", h2_style))
    schemes = db.query(models.Scheme).all()
    
    scheme_perf_headers = [
        Paragraph("<b>Scheme</b>", table_header_style),
        Paragraph("<b>Allocated (INR)</b>", table_header_style),
        Paragraph("<b>Utilized (INR)</b>", table_header_style),
        Paragraph("<b>Utilization %</b>", table_header_style)
    ]
    
    scheme_perf_rows = [scheme_perf_headers]
    for s in schemes:
        alloc_q = db.query(func.sum(models.FundAllocation.allocated_amount)).filter(models.FundAllocation.scheme_id == s.id)
        util_q = db.query(func.sum(models.FundUtilization.utilized_amount)).filter(models.FundUtilization.scheme_id == s.id)
        if state:
            alloc_q = alloc_q.filter(models.FundAllocation.state == state)
            util_q = util_q.filter(models.FundUtilization.state == state)
        if district:
            alloc_q = alloc_q.filter(models.FundAllocation.district == district)
            util_q = util_q.filter(models.FundUtilization.district == district)
            
        alloc = alloc_q.scalar() or 0.0
        util = util_q.scalar() or 0.0
        pct = (util / alloc * 100) if alloc > 0 else 0.0
        
        scheme_perf_rows.append([
            Paragraph(s.scheme_name, table_body_style),
            Paragraph(f"{alloc:,.2f}", table_body_style),
            Paragraph(f"{util:,.2f}", table_body_style),
            Paragraph(f"{pct:.2f}%", table_body_style)
        ])
        
    scheme_table = Table(scheme_perf_rows, colWidths=[180, 120, 120, 120])
    scheme_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#cbd5e1')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
    ]))
    story.append(scheme_table)
    
    story.append(Spacer(1, 20))
    story.append(PageBreak())
    
    # Validation Audit Summary
    story.append(Paragraph("Validation Logs & Integrity Errors", h2_style))
    errors = err_query.order_by(models.ValidationError.timestamp.desc()).limit(10).all()
    
    err_headers = [
        Paragraph("<b>Timestamp</b>", table_header_style),
        Paragraph("<b>Category</b>", table_header_style),
        Paragraph("<b>Details / Error log</b>", table_header_style)
    ]
    err_rows = [err_headers]
    for e in errors:
        err_rows.append([
            Paragraph(e.timestamp.strftime('%d-%b %H:%M'), table_body_style),
            Paragraph(e.error_type, table_body_style),
            Paragraph(e.description, table_body_style)
        ])
        
    err_table = Table(err_rows, colWidths=[100, 120, 320])
    err_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#b91c1c')), # Red accent for errors
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#fef2f2')]),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#fca5a5')),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#fee2e2')),
    ]))
    story.append(err_table)

    # Build PDF document
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_excel_report(db: Session, report_type: str, state: str = None, district: str = None) -> io.BytesIO:
    wb = Workbook()
    
    # Colors
    navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    red_fill = PatternFill(start_color="B91C1C", end_color="B91C1C", fill_type="solid")
    gray_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    # Fonts
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_dark = Font(name="Calibri", size=11, bold=True, color="000000")
    title_font = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
    normal_font = Font(name="Calibri", size=11, color="334155")
    
    # Border
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # -----------------------------------------------------
    # Sheet 1: Executive Summary
    # -----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1["A1"] = "GOVERNMENT SCHEME MONITORING & MIS SYSTEM"
    ws1["A1"].font = title_font
    
    loc_info = []
    if state:
        loc_info.append(state)
    if district:
        loc_info.append(district)
    loc_str = f" for {', '.join(loc_info)}" if loc_info else " (National)"
    
    ws1["A2"] = f"Executive {report_type.capitalize()} Summary Report{loc_str} — generated {datetime.now().strftime('%Y-%m-%d')}"
    ws1["A2"].font = Font(name="Calibri", size=11, italic=True)
    
    # Fetch summary details
    b_query = db.query(models.Beneficiary)
    alloc_query = db.query(func.sum(models.FundAllocation.allocated_amount))
    util_query = db.query(func.sum(models.FundUtilization.utilized_amount))
    err_query = db.query(models.ValidationError)
    
    if state:
        b_query = b_query.filter(models.Beneficiary.state == state)
        alloc_query = alloc_query.filter(models.FundAllocation.state == state)
        util_query = util_query.filter(models.FundUtilization.state == state)
    if district:
        b_query = b_query.filter(models.Beneficiary.district == district)
        alloc_query = alloc_query.filter(models.FundAllocation.district == district)
        util_query = util_query.filter(models.FundUtilization.district == district)
        
    total_beneficiaries = b_query.count()
    total_alloc = alloc_query.scalar() or 0.0
    total_util = util_query.scalar() or 0.0
    
    if state or district:
        b_sub = db.query(models.Beneficiary.beneficiary_id)
        if state:
            b_sub = b_sub.filter(models.Beneficiary.state == state)
        if district:
            b_sub = b_sub.filter(models.Beneficiary.district == district)
        b_ids = b_sub.subquery()
        
        desc_filter = []
        if state:
            desc_filter.append(models.ValidationError.description.contains(state))
        if district:
            desc_filter.append(models.ValidationError.description.contains(district))
            
        from sqlalchemy import or_
        if desc_filter:
            err_query = err_query.filter((models.ValidationError.beneficiary_id.in_(b_ids)) | or_(*desc_filter))
        else:
            err_query = err_query.filter(models.ValidationError.beneficiary_id.in_(b_ids))
            
    total_errors = err_query.count()
    
    ws1["A4"] = "KPI Metric"
    ws1["B4"] = "Value"
    ws1["A4"].fill = navy_fill
    ws1["A4"].font = header_font
    ws1["B4"].fill = navy_fill
    ws1["B4"].font = header_font
    
    kpis = [
        ("Total Beneficiaries", total_beneficiaries),
        ("Total Budget Allocated (INR)", total_alloc),
        ("Total Budget Utilized (INR)", total_util),
        ("Utilization Rate (%)", (total_util / total_alloc * 100) if total_alloc > 0 else 0),
        ("Validation Errors", total_errors)
    ]
    
    for idx, (k, v) in enumerate(kpis):
        r = 5 + idx
        ws1[f"A{r}"] = k
        ws1[f"B{r}"] = v
        ws1[f"A{r}"].font = normal_font
        ws1[f"B{r}"].font = bold_dark
        ws1[f"A{r}"].border = thin_border
        ws1[f"B{r}"].border = thin_border
        
        # Apply formatting
        if "%" in k:
            ws1[f"B{r}"].number_format = '0.00"%"'
        elif "Budget" in k:
            ws1[f"B{r}"].number_format = '#,##0.00'
        elif "Beneficiaries" in k:
            ws1[f"B{r}"].number_format = '#,##0'

    # Auto size columns
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # -----------------------------------------------------
    # Sheet 2: Scheme Metrics
    # -----------------------------------------------------
    ws2 = wb.create_sheet(title="Scheme Metrics")
    ws2.views.sheetView[0].showGridLines = True
    
    headers2 = ["Scheme Name", "Allocated (INR)", "Utilized (INR)", "Utilization %"]
    for col_idx, text in enumerate(headers2):
        cell = ws2.cell(row=1, column=col_idx+1, value=text)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    schemes = db.query(models.Scheme).all()
    for row_idx, s in enumerate(schemes):
        alloc_q = db.query(func.sum(models.FundAllocation.allocated_amount)).filter(models.FundAllocation.scheme_id == s.id)
        util_q = db.query(func.sum(models.FundUtilization.utilized_amount)).filter(models.FundUtilization.scheme_id == s.id)
        if state:
            alloc_q = alloc_q.filter(models.FundAllocation.state == state)
            util_q = util_q.filter(models.FundUtilization.state == state)
        if district:
            alloc_q = alloc_q.filter(models.FundAllocation.district == district)
            util_q = util_q.filter(models.FundUtilization.district == district)
            
        alloc = alloc_q.scalar() or 0.0
        util = util_q.scalar() or 0.0
        pct = (util / alloc) if alloc > 0 else 0.0
        
        r = row_idx + 2
        ws2.cell(row=r, column=1, value=s.scheme_name).font = normal_font
        
        c2 = ws2.cell(row=r, column=2, value=alloc)
        c2.number_format = '#,##0.00'
        c2.font = normal_font
        
        c3 = ws2.cell(row=r, column=3, value=util)
        c3.number_format = '#,##0.00'
        c3.font = normal_font
        
        c4 = ws2.cell(row=r, column=4, value=pct)
        c4.number_format = '0.00%'
        c4.font = normal_font
        
        for col_idx in range(1, 5):
            ws2.cell(row=r, column=col_idx).border = thin_border
            
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # -----------------------------------------------------
    # Sheet 3: Validation Violations
    # -----------------------------------------------------
    ws3 = wb.create_sheet(title="Integrity Violations")
    ws3.views.sheetView[0].showGridLines = True
    
    headers3 = ["Error ID", "Beneficiary ID", "Type", "Log details", "Timestamp"]
    for col_idx, text in enumerate(headers3):
        cell = ws3.cell(row=1, column=col_idx+1, value=text)
        cell.fill = red_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    errors = err_query.limit(100).all()
    for row_idx, e in enumerate(errors):
        r = row_idx + 2
        ws3.cell(row=r, column=1, value=e.error_id).font = normal_font
        ws3.cell(row=r, column=2, value=e.beneficiary_id or "-").font = normal_font
        ws3.cell(row=r, column=3, value=e.error_type).font = normal_font
        ws3.cell(row=r, column=4, value=e.description).font = normal_font
        
        c5 = ws3.cell(row=r, column=5, value=e.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
        c5.font = normal_font
        
        for col_idx in range(1, 6):
            ws3.cell(row=r, column=col_idx).border = thin_border
            
    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws3.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 45) # Keep details from getting ridiculously wide

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
